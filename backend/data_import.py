from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from backend.database import Database


def _clean(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.isoformat()
    return value


def _table(sheet: Any) -> list[dict[str, Any]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() for value in rows[0] if value is not None]
    output: list[dict[str, Any]] = []
    for row in rows[1:]:
        values = list(row[: len(headers)])
        if not any(value is not None for value in values):
            continue
        output.append({header: _clean(value) for header, value in zip(headers, values, strict=True)})
    return output


def _read_snapshot(readme_sheet: Any) -> str:
    for row in readme_sheet.iter_rows(values_only=True):
        if row and str(row[0]).strip().lower() == "dataset snapshot":
            return str(row[1]).strip()
    raise ValueError("Workbook README does not contain a dataset snapshot")


def import_assessment_data(db: Database, workbook_path: Path, users_seed_path: Path) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    snapshot = _read_snapshot(workbook["README"])
    accounts = _table(workbook["accounts"])
    orders = _table(workbook["orders"])
    tickets = _table(workbook["tickets"])
    users = json.loads(users_seed_path.read_text(encoding="utf-8"))

    db.initialize()
    with db.transaction() as connection:
        for table in ("orders", "tickets", "accounts", "customer_users"):
            connection.execute(f"DELETE FROM {table}")  # noqa: S608 - fixed internal table list
        connection.execute(
            "INSERT OR REPLACE INTO metadata (key, value) VALUES ('dataset_snapshot', ?)",
            (snapshot,),
        )

        connection.executemany(
            """INSERT INTO accounts
            (account_id, account_name, plan, status, csm, contract_file, premium_support, notes)
            VALUES (:account_id, :account_name, :plan, :status, :csm, :contract_file, :premium_support, :notes)""",
            [row | {"premium_support": int(bool(row["premium_support"]))} for row in accounts],
        )
        connection.executemany(
            """INSERT INTO orders
            (order_id, account_id, carrier, status, booked_at, pickup_window_start, pickup_window_end,
             pickup_actual_at, shipment_fee_inr, carrier_fault, customer_fault, cancellation_requested_at, notes)
            VALUES (:order_id, :account_id, :carrier, :status, :booked_at, :pickup_window_start, :pickup_window_end,
                    :pickup_actual_at, :shipment_fee_inr, :carrier_fault, :customer_fault, :cancellation_requested_at, :notes)""",
            [
                row
                | {
                    "carrier_fault": int(bool(row["carrier_fault"])),
                    "customer_fault": int(bool(row["customer_fault"])),
                }
                for row in orders
            ],
        )
        connection.executemany(
            """INSERT INTO tickets
            (ticket_id, account_id, created_at, status, subject, description, channel, assigned_to,
             last_customer_message_at, historical_resolution)
            VALUES (:ticket_id, :account_id, :created_at, :status, :subject, :description, :channel, :assigned_to,
                    :last_customer_message_at, :historical_resolution)""",
            tickets,
        )
        connection.executemany(
            """INSERT INTO customer_users
            (user_id, username, display_name, account_id, password_hash, is_active)
            VALUES (:user_id, :username, :display_name, :account_id, :password_hash, :is_active)""",
            [row | {"is_active": int(bool(row["is_active"]))} for row in users],
        )

    return {
        "dataset_snapshot": snapshot,
        "accounts": len(accounts),
        "orders": len(orders),
        "tickets": len(tickets),
        "users": len(users),
    }
